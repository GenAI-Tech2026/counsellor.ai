#!/usr/bin/env python3
"""Build tidy KCET-2024 engineering cutoff Excel files from the source PDFs."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from parse_kcet import parse_pdf

HEADERS = ['College Code', 'College Name', 'Place',
           'Branch Code', 'Branch Name', 'Category', 'Closing Rank']
WIDTHS = [12, 52, 28, 11, 30, 10, 13]

# (source pdf, region, round, sheet name, output filename)
FILES = [
    ('ENGG_CUTOFF_2024_GEN_MOCK.pdf', 'General',          'Mock',    'GEN Mock', 'KCET_2024_ENGG_GEN_Mock_CutOff.xlsx'),
    ('ENGG_CUTOFF_2024_GEN_R1.pdf',   'General',          'Round 1', 'GEN R1',   'KCET_2024_ENGG_GEN_Round1_CutOff.xlsx'),
    ('ENGG_CUTOFF_2024_GEN_R2.pdf',   'General',          'Round 2', 'GEN R2',   'KCET_2024_ENGG_GEN_Round2_CutOff.xlsx'),
    ('ENGG_CUTOFF_2024_HK_MOCK.pdf',  'Hyderabad-Karnataka', 'Mock',    'HK Mock',  'KCET_2024_ENGG_HK_Mock_CutOff.xlsx'),
    ('ENGG_CUTOFF_2024_HK_R1.pdf',    'Hyderabad-Karnataka', 'Round 1', 'HK R1',    'KCET_2024_ENGG_HK_Round1_CutOff.xlsx'),
    ('ENGG_CUTOFF_2024_HK_R2.pdf',    'Hyderabad-Karnataka', 'Round 2', 'HK R2',    'KCET_2024_ENGG_HK_Round2_CutOff.xlsx'),
]

def style_sheet(ws, nrows):
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{nrows + 1}"

def write_rows(ws, rows):
    for r in rows:
        ws.append(r[:6] + [int(r[6])])
    style_sheet(ws, len(rows))

def main():
    parsed = []
    combo = Workbook()
    combo.remove(combo.active)
    for pdf, region, rnd, sheet, out in FILES:
        meta, rows = parse_pdf(os.path.join('source', pdf))
        parsed.append((sheet, rows))
        # individual workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        write_rows(ws, rows)
        wb.save(out)
        # combined workbook sheet
        write_rows(combo.create_sheet(sheet), rows)
        print(f"{out:42s} rows={len(rows):6d}  ({meta['title']})")
    combo.save('KCET_2024_ENGG_AllRounds_CutOff.xlsx')
    total = sum(len(r) for _, r in parsed)
    print(f"KCET_2024_ENGG_AllRounds_CutOff.xlsx       sheets={len(parsed)} total_rows={total}")

if __name__ == '__main__':
    main()
