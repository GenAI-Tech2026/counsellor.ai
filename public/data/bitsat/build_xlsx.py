#!/usr/bin/env python3
"""Build the BITSAT cut-off Excel workbook from the official BITS admissions archive page.

Why this source? BITSAT cut-offs ARE published officially by BITS Pilani itself (unlike CUET,
where the NTA publishes nothing). The richest official artifact is the admissions-portal page
  https://admissions.bits-pilani.ac.in/FD/BITSAT_cutOffs.html
which embeds the FINAL cut-off scores for academic years 2017-18 .. 2025-26 across the Pilani,
K. K. Birla Goa and Hyderabad campuses (Dubai is not part of the BITSAT cut-off lists). All nine
years are present as hidden <div id="YYYY-YYYY"> blocks toggled by a ?yr= query param, so a single
saved copy of the page contains everything. (The WordPress filter tool at
bits-pilani.ac.in/cut-off-bitsat-scores only exposes 2017-2020 + 2022, so this archive is richer.)

BITSAT cut-offs are SCORES out of 390 (not ranks) — the lowest BITSAT score on which a seat was
allocated for a given campus x programme in that year's final allocation.

The page mixes two table layouts across years, both handled here:
  NEW (2021-22 .. 2025-26): one table, explicit columns  Campus | Program | Cut-off Score | Max.
  OLD (2017-18 .. 2020-21): one table PER campus; the campus name is in the column header
                            "Degree programme at <Campus> Campus", rows are Program | Score | Max.

Output: BITSAT_CutOffs_2017_2025.xlsx
  - sheet "All Years"  tidy/long: Academic Year, Campus, Programme, Cut-off Score, Max Marks
  - one sheet per academic year, same columns (minus Academic Year)

Reads values directly from the saved HTML (not retyped); re-running reproduces them exactly.
Requires beautifulsoup4 + lxml + openpyxl.
"""
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SRC = 'source/BITSAT_cutOffs_archive.html'
OUT = 'BITSAT_CutOffs_2017_2025.xlsx'

YEAR_DIV = re.compile(r'^\d{4}-\d{4}$')
CAMPUS_IN_HDR = re.compile(r'\bat\s+(.+?)\s+Campus\b', re.I)
NUM = re.compile(r'^\d+(\.\d+)?$')


def _clean(s):
    return re.sub(r'\s+', ' ', (s or '').replace('\xa0', ' ')).strip()


def _norm_campus(s):
    s = _clean(s)
    low = s.lower()
    if 'pilani' in low and 'goa' not in low:
        return 'Pilani'
    if 'goa' in low:
        return 'Goa'
    if 'hyderabad' in low or 'hyd' in low:
        return 'Hyderabad'
    if 'dubai' in low:
        return 'Dubai'
    return s


def _rows(table):
    out = []
    for tr in table.find_all('tr'):
        cells = [_clean(td.get_text(' ', strip=True)) for td in tr.find_all(['td', 'th'])]
        if any(cells):
            out.append(cells)
    return out


def parse():
    soup = BeautifulSoup(open(SRC, encoding='utf-8', errors='ignore').read(), 'lxml')
    years = {}  # 'YYYY-YY' -> list of [campus, programme, score, max]
    for div in soup.find_all('div', id=lambda v: v and YEAR_DIV.match(v)):
        yid = div['id']                       # e.g. 2025-2026
        label = f"{yid[:4]}-{yid[7:9]}"       # -> 2025-26
        recs = []
        # only LEAF tables (no nested <table>) — the page wraps content in layout tables
        for table in div.find_all('table'):
            if table.find('table'):
                continue
            rows = _rows(table)
            if not rows:
                continue
            # find the header row (mentions a cut-off / score / max-marks header)
            hidx, has_campus_col, campus_from_hdr = None, False, None
            for i, r in enumerate(rows):
                # the real column-header row has >=2 cells; the intro paragraph is a single cell
                if len(r) < 2:
                    continue
                joined = ' | '.join(r).lower()
                if 'cut-off' in joined or 'cut off' in joined or 'maximum marks' in joined:
                    hidx = i
                    has_campus_col = any(c.lower().strip() == 'campus' for c in r)
                    for c in r:
                        m = CAMPUS_IN_HDR.search(c)
                        if m:
                            campus_from_hdr = _norm_campus(m.group(1))
                    break
            if hidx is None:
                continue
            for r in rows[hidx + 1:]:
                if has_campus_col:
                    # Campus | Program | Score | Max  (Max optional)
                    if len(r) < 3:
                        continue
                    campus, prog, score = r[0], r[1], r[2]
                    mx = r[3] if len(r) > 3 else '390'
                else:
                    # Program | Score | Max ; campus from header
                    if len(r) < 2:
                        continue
                    campus, prog, score = campus_from_hdr, r[0], r[1]
                    mx = r[2] if len(r) > 2 else '390'
                if not prog or not NUM.match(score):
                    continue
                if 'programme' in prog.lower() or 'program' == prog.lower():
                    continue
                recs.append([_norm_campus(campus), _clean(prog),
                             float(score) if '.' in score else int(score),
                             int(mx) if NUM.match(mx) else 390])
        if recs:
            years[label] = recs
    return years


def _hdr(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def _finalize(ws, ncols, nrows):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def main():
    years = parse()
    order = sorted(years, key=lambda y: int(y[:4]))
    wb = Workbook()

    ws = wb.active
    ws.title = 'All Years'
    _hdr(ws, ['Academic Year', 'Campus', 'Programme', 'Cut-off Score', 'Max Marks'],
         [16, 14, 46, 14, 12])
    n = 0
    for y in order:
        for campus, prog, score, mx in years[y]:
            ws.append([y, campus, prog, score, mx])
            n += 1
    _finalize(ws, 5, n)

    for y in order:
        sh = wb.create_sheet(y)
        _hdr(sh, ['Campus', 'Programme', 'Cut-off Score', 'Max Marks'], [14, 46, 14, 12])
        for campus, prog, score, mx in years[y]:
            sh.append([campus, prog, score, mx])
        _finalize(sh, 4, len(years[y]))

    wb.save(OUT)
    print(f'{OUT}: {n} rows across {len(order)} years')
    for y in order:
        camps = {}
        for c, *_ in years[y]:
            camps[c] = camps.get(c, 0) + 1
        print(f'  {y}: {len(years[y]):3d} rows  ' +
              ' '.join(f'{c}={camps[c]}' for c in sorted(camps)))


if __name__ == '__main__':
    main()
