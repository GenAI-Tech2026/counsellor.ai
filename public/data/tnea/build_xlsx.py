#!/usr/bin/env python3
"""
Build the TNEA mark-cutoff .xlsx from the official DoTE (Anna University) PDFs.

Input : source/{BArch,Vocational}_{2021..2024}_Mark_Cutoff.pdf
Output: TNEA_MarkCutoffs_2021_2024.xlsx   (one tidy/long sheet)

TNEA (Tamil Nadu Engineering Admissions) publishes, per year, a "MARK CUTOFF"
PDF per stream. Each is one table:
    COLLEGE CODE | COLLEGE NAME | BRANCH CODE | BRANCH NAME | <category columns...>
The category columns are the *last allotted mark* (cutoff) for each TN communal
category. The cutoff mark is out of 200 for the regular streams and out of 400
for B.Arch (it folds in the NATA/aptitude component) — values are transcribed
exactly as printed, so each stream keeps its own scale.

The category set is NOT constant across years:
  - 2022-2024: OC, BC, BCM, MBC, SC, SCA, ST                       (7 categories)
  - 2021     : OC, BC, BCM, MBCV, MBC DNC, MBC, SC, SCA, ST        (9 categories)
    (in 2021 the Most Backward Classes block is itemised into Vanniyakula
     Kshatriya (MBCV), the Denotified Communities (MBC DNC) and the residual MBC.)

A tidy / long layout (one row per College x Branch x Category) is used instead of
one column per category — mirroring the kcet/, mhtcet/ and wbjee/ datasets. This
reproduces the source exactly: a year only emits the category codes it actually
prints, with no blank-cell guesswork and no fabricated zeros.

Re-run:  python3 build_xlsx.py   (from this directory)
"""
import glob
import os
import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = "TNEA_MarkCutoffs_2021_2024.xlsx"

# Canonicalise the category header text (the PDF text layer is noisy) -> clean code.
CATEGORY_CANON = {
    "OC": "OC",
    "BC": "BC",
    "BCM": "BCM",
    "MBC": "MBC",
    "MBCV": "MBCV",
    "MBCDNC": "MBC DNC",
    "MBC DNC": "MBC DNC",
    "SC": "SC",
    "SCA": "SCA",
    "ST": "ST",
}

# Stable display order for categories within a (college, branch) group.
CAT_ORDER = ["OC", "BC", "BCM", "MBCV", "MBC DNC", "MBC", "SC", "SCA", "ST"]


def norm(s):
    """Collapse the embedded line wraps pdfplumber keeps in a cell."""
    return re.sub(r"\s+", " ", str(s or "")).strip()


def is_header(row):
    return bool(row) and norm(row[0]).upper().startswith("COLLEGE")


def category_columns(header):
    """Map column index (>=4) -> canonical category code from a header row."""
    cols = {}
    for i in range(4, len(header)):
        key = norm(header[i]).upper().replace(" ", "")
        code = CATEGORY_CANON.get(key) or CATEGORY_CANON.get(norm(header[i]).upper())
        if code:
            cols[i] = code
    return cols


def parse_mark(v):
    """'332.3333' -> 332.3333, integers stay int, blanks/garbage -> None."""
    s = norm(v)
    if not s:
        return None
    s = s.replace(",", "")
    if not re.fullmatch(r"\d+(\.\d+)?", s):
        return None
    f = float(s)
    return int(f) if f.is_integer() else round(f, 4)


def parse_pdf(path):
    """Yield tidy records from one stream-year PDF."""
    base = os.path.basename(path)
    m = re.match(r"(BArch|Vocational)_(\d{4})_", base)
    stream, year = m.group(1), int(m.group(2))

    catmap = {}  # carried across pages; the header repeats but guard anyway
    records = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if is_header(row):
                        catmap = category_columns(row)
                        continue
                    if not catmap:
                        continue
                    code = norm(row[0])
                    if not code:  # continuation fragment with no college code
                        continue
                    name = norm(row[1])
                    bcode = norm(row[2])
                    bname = norm(row[3])
                    for i, cat in catmap.items():
                        if i >= len(row):
                            continue
                        mark = parse_mark(row[i])
                        if mark is None:
                            continue
                        records.append({
                            "year": year, "stream": stream,
                            "college_code": code, "college_name": name,
                            "branch_code": bcode, "branch_name": bname,
                            "category": cat, "cutoff": mark,
                        })
    return records


def write_sheet(ws, records):
    headers = ["Year", "Stream", "College Code", "College Name",
               "Branch Code", "Branch Name", "Category", "Cutoff Mark"]
    widths = [7, 12, 13, 52, 13, 40, 10, 12]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    def code_sort(c):
        return (0, int(c)) if c.isdigit() else (1, c)

    records = sorted(records, key=lambda r: (
        r["stream"], r["year"], code_sort(r["college_code"]),
        r["branch_code"],
        CAT_ORDER.index(r["category"]) if r["category"] in CAT_ORDER else 99,
    ))
    for r in records:
        ws.append([r["year"], r["stream"], r["college_code"], r["college_name"],
                   r["branch_code"], r["branch_name"], r["category"], r["cutoff"]])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
    return len(records)


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    records = []
    for path in sorted(glob.glob("source/*_Mark_Cutoff.pdf")):
        recs = parse_pdf(path)
        print(f"  {os.path.basename(path):34s} -> {len(recs):5d} tidy rows")
        records.extend(recs)

    wb = Workbook()
    n = write_sheet(wb.active, records)
    wb.active.title = "Mark Cutoffs"
    wb.save(OUT)

    streams = sorted({r["stream"] for r in records})
    print(f"\n{OUT}: {n} tidy rows  (streams: {', '.join(streams)})")


if __name__ == "__main__":
    main()
