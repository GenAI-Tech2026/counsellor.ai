#!/usr/bin/env python3
"""Build the CUET cut-off Excel workbook from the official DU CSAS UG 2025 PDFs in ./source.

Why DU CSAS? The NTA does not publish any CUET cut-offs — CUET only yields a score, and
cut-offs are set by each participating university. The University of Delhi (the largest CUET
university) publishes the most comprehensive official cut-offs: round-wise "Minimum Allocation
Score" lists, i.e. the lowest CUET (UG) score on which a seat was allocated for every
college x programme x category. These ARE the CUET cut-offs in practice.

Source (admission.uod.ac.in, Undergraduate Admissions 2025-26):
  DU_CSAS_UG_2025_Round1_CutOff.pdf    First round  — wide layout, all six categories
  DU_CSAS_UG_2025_Round3_CutOff.pdf    Third round  — wide layout (blank where no allocation)
  DU_CSAS_UG_2025_SpotRound_CutOff.pdf Spot round   — tidy layout, one category per row

Scores are out of 1000 (CUET-UG normalized total). The values are read directly from the PDFs
by word x-position (page width ~842), so re-running reproduces them exactly.

Layouts:
  WIDE: S.NO | COLLEGE | PROGRAM | UR OBC SC ST EWS PwBD
  SPOT: S.NO | PROGRAM | COLLEGE | CATEGORY | SCORE
Records are keyed off the integer in the S.NO column (so fully-blank rows in later rounds are
kept); wrapped programme names spill onto the line ABOVE the S.No line and are buffered;
centred page-number footers carry no S.No/text and are dropped.
"""
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

CATS = ['UR', 'OBC', 'SC', 'ST', 'EWS', 'PwBD']
INT = re.compile(r'^\d+$')
NUM = re.compile(r'^\d+(\.\d+)?$')

R1 = 'source/DU_CSAS_UG_2025_Round1_CutOff.pdf'
R3 = 'source/DU_CSAS_UG_2025_Round3_CutOff.pdf'
RS = 'source/DU_CSAS_UG_2025_SpotRound_CutOff.pdf'
OUT = 'DU_CSAS_UG_2025_CUET_CutOffs.xlsx'


def _lines(page):
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    buckets = {}
    for w in words:
        buckets.setdefault(round(w['top'] / 4), []).append(w)
    return [sorted(buckets[k], key=lambda w: w['x0']) for k in sorted(buckets)]


def _header_top(lines, needle):
    for ln in lines:
        t = ' '.join(w['text'] for w in ln).upper()
        if 'S.NO' in t and needle in t:
            return ln[0]['top']
    return None


def _cat_of(center):
    for i, b in enumerate((482, 550, 618, 686, 754)):
        if center < b:
            return i
    return 5


def parse_wide(path):
    """-> [college, program, UR, OBC, SC, ST, EWS, PwBD]"""
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines = _lines(page)
            htop = _header_top(lines, 'COLLEGE')
            pend_col, pend_prog, cur = [], [], None
            for ln in lines:
                if htop is not None and ln[0]['top'] <= htop + 2:
                    continue
                sno, col_w, prog_w, scores = None, [], [], [None] * 6
                for w in ln:
                    x0, txt = w['x0'], w['text']
                    if x0 < 60:
                        if INT.match(txt):
                            sno = txt
                    elif x0 < 215:
                        col_w.append(txt)
                    elif x0 < 410:
                        prog_w.append(txt)
                    elif NUM.match(txt):
                        scores[_cat_of((x0 + w['x1']) / 2)] = float(txt)
                if sno is not None:
                    if cur:
                        rows.append(cur)
                    cur = [' '.join(pend_col + col_w).strip(),
                           ' '.join(pend_prog + prog_w).strip()] + scores
                    pend_col, pend_prog = [], []
                else:
                    pend_col += col_w
                    pend_prog += prog_w
            if cur:
                rows.append(cur)
    return [r for r in rows if r[0] or r[1]]


def parse_spot(path):
    """-> [college, program, category, score]"""
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines = _lines(page)
            htop = _header_top(lines, 'PROGRAM')
            pend_prog, cur = [], None
            for ln in lines:
                if htop is not None and ln[0]['top'] <= htop + 2:
                    continue
                sno, prog_w, col_w, cat, score = None, [], [], '', None
                for w in ln:
                    x0, txt = w['x0'], w['text']
                    if x0 < 180:
                        if INT.match(txt):
                            sno = txt
                    elif x0 < 365:
                        prog_w.append(txt)
                    elif x0 < 530:
                        col_w.append(txt)
                    elif x0 < 595:
                        cat = txt
                    elif NUM.match(txt):
                        score = float(txt)
                if sno is not None:
                    if cur:
                        rows.append(cur)
                    cur = [' '.join(col_w).strip(),
                           ' '.join(pend_prog + prog_w).strip(), cat, score]
                    pend_prog = []
                else:
                    pend_prog += prog_w
            if cur:
                rows.append(cur)
    return [r for r in rows if r[0] or r[1]]


def write_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def finalize(ws, ncols, nrows):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def main():
    wb = Workbook()

    # Round 1 (wide)
    ws = wb.active
    ws.title = 'Round 1'
    write_header(ws, ['College', 'Programme'] + [f'{c} cut-off' for c in CATS],
                 [42, 52, 12, 12, 12, 12, 12, 12])
    r1 = parse_wide(R1)
    for row in r1:
        ws.append(row)
    finalize(ws, 8, len(r1))

    # Round 3 (wide)
    ws3 = wb.create_sheet('Round 3')
    write_header(ws3, ['College', 'Programme'] + [f'{c} cut-off' for c in CATS],
                 [42, 52, 12, 12, 12, 12, 12, 12])
    r3 = parse_wide(R3)
    for row in r3:
        ws3.append(row)
    finalize(ws3, 8, len(r3))

    # Spot round (tidy)
    wss = wb.create_sheet('Spot Round')
    write_header(wss, ['College', 'Programme', 'Category', 'Cut-off Score'],
                 [42, 52, 12, 16])
    rs = parse_spot(RS)
    for row in rs:
        wss.append(row)
    finalize(wss, 4, len(rs))

    wb.save(OUT)
    print(f'{OUT}: Round 1={len(r1)}, Round 3={len(r3)}, Spot Round={len(rs)} rows '
          f'({len(set(r[0] for r in r1))} colleges)')


if __name__ == '__main__':
    main()
