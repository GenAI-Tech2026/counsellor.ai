#!/usr/bin/env python3
"""
Build the KEAM 2025 Engineering cut-off .xlsx files from the official CEE Kerala
"Last Rank" PDFs.

Input : source/KEAM_2025_ENGG_Phase1_LastRank.pdf  (see source/fetch_lastrank.py)
        source/KEAM_2025_ENGG_Phase2_LastRank.pdf
Output: KEAM_2025_ENGG_Phase1_LastRank.xlsx
        KEAM_2025_ENGG_Phase2_LastRank.xlsx
        KEAM_2025_ENGG_AllPhases_LastRank.xlsx   (one sheet per phase)

Tidy / long layout, mirroring the kcet/, mhtcet/ and wbjee/ datasets: one row per
Branch x College x Category, carrying the official Last Rank for that phase.

Re-run:  python3 build_xlsx.py   (from this directory)
"""
import os
import sys
import importlib.util

from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source")

# import the word-position PDF parser kept next to the source PDFs
spec = importlib.util.spec_from_file_location("parse_pdf", os.path.join(SRC, "parse_pdf.py"))
parse_pdf = importlib.util.module_from_spec(spec)
spec.loader.exec_module(parse_pdf)

PHASES = [
    ("Phase 1", "KEAM_2025_ENGG_Phase1_LastRank.pdf", "KEAM_2025_ENGG_Phase1_LastRank.xlsx"),
    ("Phase 2", "KEAM_2025_ENGG_Phase2_LastRank.pdf", "KEAM_2025_ENGG_Phase2_LastRank.xlsx"),
]

HEADERS = ["Branch", "College Code", "College", "College Type", "Category", "Last Rank"]

TYPE_LABEL = {"G": "Government", "S": "Self Financing", "N": "Government Controlled"}

# canonical KEAM category order (the 13 reservation columns), then special codes A-Z
MAIN_ORDER = ["SM", "EZ", "MU", "LA", "DV", "VK", "BH", "BX", "KN", "KU", "SC", "ST", "EW"]
ORDER_IDX = {c: i for i, c in enumerate(MAIN_ORDER)}


def cat_key(cat):
    return (0, ORDER_IDX[cat]) if cat in ORDER_IDX else (1, cat)


def tidy_rows(pdf_path):
    rows = parse_pdf.parse(pdf_path)
    rows.sort(key=lambda r: (r[0] or "", r[2] or "", r[1] or "", cat_key(r[4])))
    out = []
    for branch, code, college, typ, cat, rank in rows:
        out.append([branch, code, college, TYPE_LABEL.get(typ, typ), cat, rank])
    return out


def write_sheet(ws, rows):
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    # widths for readability
    for col, w in zip("ABCDEF", (46, 12, 52, 22, 10, 11)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def main():
    all_wb = Workbook()
    all_wb.remove(all_wb.active)

    for label, pdf_name, xlsx_name in PHASES:
        pdf_path = os.path.join(SRC, pdf_name)
        rows = tidy_rows(pdf_path)

        wb = Workbook()
        write_sheet(wb.active, rows)
        wb.active.title = label
        wb.save(os.path.join(HERE, xlsx_name))

        write_sheet(all_wb.create_sheet(label), rows)

        colleges = len({(r[1], r[2]) for r in rows})
        branches = len({r[0] for r in rows})
        print(f"{xlsx_name:42} {len(rows):5} rows  {branches:3} branches  {colleges:3} colleges")

    all_path = os.path.join(HERE, "KEAM_2025_ENGG_AllPhases_LastRank.xlsx")
    all_wb.save(all_path)
    print(f"{'KEAM_2025_ENGG_AllPhases_LastRank.xlsx':42} (one sheet per phase)")


if __name__ == "__main__":
    sys.exit(main())
