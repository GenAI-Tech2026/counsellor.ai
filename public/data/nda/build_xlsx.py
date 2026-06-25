#!/usr/bin/env python3
"""Build the NDA & NA cut-off Excel workbook (2020-2024) from the official UPSC PDFs.

Each source PDF holds a single small "Recommendation Details and Cut-off Marks" table.
The written-stage cut-off (out of 900) and final-stage cut-off (out of 1800) are exam-wide
single values; vacancies are given per wing. Values below were transcribed by column position
from the source PDFs in ./source and validated cell-for-cell against them.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# One record per examination (year x session).
#   written  = minimum qualifying marks at the written stage (out of 900)
#   subj_min = minimum % required in each subject at the written stage
#   final    = marks of the last recommended candidate at the final stage (out of 1800)
#   vacancies / recommended = totals for the exam (None where the source omits them)
CUTOFFS = [
    # year, session, written, subj_min%, final, total_vacancies, recommended
    (2020, 'I',  355, 25, 723, 418, 533),
    (2020, 'II', 355, 25, 719, None, None),   # older format: only the two cut-off figures
    (2021, 'I',  343, 25, 709, 400, 517),
    (2021, 'II', 355, 25, 726, 400, 462),
    (2022, 'I',  360, 25, 720, 400, 519),
    (2022, 'II', 316, 20, 678, 400, 538),
    (2023, 'I',  301, 25, 664, 395, 628),
    (2023, 'II', 292, 20, 656, 395, 699),
    (2024, 'I',  291, 20, 654, 400, 641),
    (2024, 'II', 305, 25, 673, 404, 792),
]

# Vacancies per wing, kept exactly as the source PDFs list them — including the Air Force's
# Flying / Ground Duties (Tech) / Ground Duties (Non-Tech) sub-branches.
# Each row: (wing, sub_wing, vacancies, female, notes).
#   sub_wing = '' for wings the source does not subdivide.
#   female   = vacancies reserved for female candidates (None where the source omits the split).
# Air Force totals 120 every year (92 + 18 + 10). The 2020 II PDF gives no wing breakdown.
WINGS = {
    (2020, 'I'): [
        ('Army', '', 208, None, ''), ('Navy', '', 42, None, ''),
        ('Air Force', '', 120, None, 'incl. 28 ground duties'), ('Naval Academy', '', 48, None, ''),
    ],
    (2021, 'I'): [
        ('Army', '', 208, None, ''), ('Navy', '', 42, None, ''),
        ('Air Force', '', 120, None, 'incl. 28 ground duties'), ('Naval Academy', '', 30, None, ''),
    ],
    (2021, 'II'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 3, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 30, 0, 'male candidates only'),
    ],
    (2022, 'I'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 3, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 30, 0, 'male candidates only'),
    ],
    (2022, 'II'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 3, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 30, 0, 'male candidates only'),
    ],
    (2023, 'I'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 3, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 25, 0, 'male candidates only'),
    ],
    (2023, 'II'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 12, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 25, 7, ''),
    ],
    (2024, 'I'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 6, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 30, 9, ''),
    ],
    (2024, 'II'): [
        ('Army', '', 208, 10, ''), ('Navy', '', 42, 6, ''),
        ('Air Force', 'Flying', 92, 2, ''), ('Air Force', 'Ground Duties (Tech)', 18, 2, ''),
        ('Air Force', 'Ground Duties (Non-Tech)', 10, 2, ''),
        ('Naval Academy', '', 34, 5, '10+2 Cadet Entry Scheme'),
    ],
}

OUT = 'NDA_NA_CutOffs_2020_2024.xlsx'


def write_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def finalize(ws, ncols, nrows):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def main():
    wb = Workbook()

    # Sheet 1 — headline cut-offs (one row per exam)
    ws = wb.active
    ws.title = 'Cut-offs'
    headers = ['Year', 'Session', 'Examination',
               'Written Cut-off (out of 900)', 'Min % per Subject (written)',
               'Final Cut-off (out of 1800)', 'Total Vacancies', 'Candidates Recommended']
    write_header(ws, headers, [7, 9, 22, 17, 16, 16, 14, 16])
    for y, s, written, subj, final, vac, rec in CUTOFFS:
        ws.append([y, s, f'NDA & NA ({s}), {y}', written, subj, final, vac, rec])
    finalize(ws, len(headers), len(CUTOFFS))

    # Sheet 2 — vacancies by wing & sub-wing (tidy / long)
    ws2 = wb.create_sheet('Vacancies by Wing')
    headers2 = ['Year', 'Session', 'Wing', 'Sub-Wing', 'Vacancies', 'Female Vacancies', 'Notes']
    write_header(ws2, headers2, [7, 9, 16, 26, 12, 16, 24])
    n = 0
    for (y, s), rows in sorted(WINGS.items()):
        for wing, sub, vac, fem, notes in rows:
            ws2.append([y, s, wing, sub, vac, fem, notes])
            n += 1
    finalize(ws2, len(headers2), n)

    wb.save(OUT)
    print(f'{OUT}: Cut-offs={len(CUTOFFS)} rows, Vacancies by Wing={n} rows')


if __name__ == '__main__':
    main()
