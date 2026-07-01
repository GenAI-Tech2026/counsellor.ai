#!/usr/bin/env python3
"""Build the COMEDK UGET engineering cut-off Excel workbook from the official COMEDK PDFs.

Why this source? COMEDK (Consortium of Medical, Engineering and Dental Colleges of Karnataka)
publishes its UGET counselling cut-off RANKS officially on its own portal, comedk.org. After each
counselling cycle it issues an "Engineering cut-off after all rounds" PDF — the FINAL closing rank
on which a seat was allotted for each college x branch x seat-category. Those final PDFs are the
single most complete official artifact (they fold in every round), so this dataset is built from
them. The verbatim PDFs are preserved under source/:
  - source/COMEDK_2024_Engineering_Cutoff_AfterAllRounds.pdf
      https://www.comedk.org/uploads/Engineering-2024-Cut-off-Ranks-after-all-rounds-Notified-on-17_07_2025.pdf
  - source/COMEDK_2023_Engineering_Cutoff_AfterAllRounds.pdf
      https://www.comedk.org/uploads/2023-Engineering-Cut-Off-Ranks-After-All-Rounds-Notified-on-27_05_2024-final.pdf

A COMEDK cut-off is a RANK (the candidate's COMEDK UGET rank), like KCET/WBJEE/KEAM — NOT a score.
"Closing Rank" is the last (highest-numbered) rank allotted to that college-branch-category.

PDF layout: a wide matrix — rows are (College Code, College Name, Seat Category), and there is one
column per branch. The matrix is too wide for one page, so the branch columns are split into groups
of ten and every group repeats the full college list across ~13 pages. pdfplumber extracts each
page's table cleanly; this script melts every page into tidy/long rows and drops empty cells (a
college-branch-category with no allotment is simply absent from the source, so it is absent here —
no blank-cell guesswork). Seat categories present: GM (General Merit) and KKR (Kalyana-Karnataka /
Article 371(J) local-reservation seats, formerly Hyderabad-Karnataka).

Output: COMEDK_Engineering_CutOffs_2023_2024.xlsx
  - sheet "All Years"  tidy/long: Year, College Code, College Name, Seat Category,
                                  Branch Code, Branch Name, Closing Rank
  - one sheet per year (2023, 2024): same columns minus Year

Values are read directly from the saved PDFs (not retyped); re-running reproduces them exactly.
Requires pdfplumber + openpyxl.  Run:  python3 build_xlsx.py   (from this directory)
"""
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pdfplumber

CODE_RE = re.compile(r'^E\d+', re.I)
RANK_RE = re.compile(r'^\d+$')
YEAR_RE = re.compile(r'(20\d{2})')


def _norm(s):
    return re.sub(r'\s+', ' ', (s or '').replace('\xa0', ' ')).strip()


def _branch(header):
    """'AD-Artificial Intelligence & Data Science' -> ('AD', 'Artificial Intelligence & Data Science')
    Tolerates spacing variants: 'MAE- MECHANICAL...', 'IAR - Information...', 'IDA -Information...'."""
    h = _norm(header)
    m = re.match(r'^([A-Za-z]+)\s*-\s*(.+)$', h)
    if not m:
        return h, h
    return m.group(1).upper(), _norm(m.group(2))


def parse_pdf(path):
    """Return list of [code, name, category, branch_code, branch_name, closing_rank]."""
    rows = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if not tables:
                continue
            tbl = tables[0]
            if not tbl or len(tbl) < 2:
                continue
            branch_cols = [_branch(c) for c in tbl[0][3:]]
            for r in tbl[1:]:
                if not r or not r[0]:
                    continue
                code = _norm(r[0])
                if not CODE_RE.match(code):
                    continue
                name = _norm(r[1])
                category = _norm(r[2])
                for j, (bc, bn) in enumerate(branch_cols):
                    cell = _norm(r[3 + j]) if 3 + j < len(r) else ''
                    if cell and RANK_RE.match(cell):
                        rows.append([code, name, category, bc, bn, int(cell)])
    # stable order: college code, branch code, category
    rows.sort(key=lambda x: (x[0], x[3], x[2]))
    return rows


def _hdr(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def _finalize(ws, ncols, nrows):
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


def main():
    files = {}
    for p in sorted(glob.glob('source/*.pdf')):
        m = YEAR_RE.search(p)
        if m:
            files[m.group(1)] = p
    years = sorted(files)

    data = {y: parse_pdf(files[y]) for y in years}

    wb = Workbook()
    ws = wb.active
    ws.title = 'All Years'
    _hdr(ws, ['Year', 'College Code', 'College Name', 'Seat Category',
              'Branch Code', 'Branch Name', 'Closing Rank'],
         [8, 12, 52, 12, 12, 52, 13])
    n = 0
    for y in years:
        for code, name, cat, bc, bn, rank in data[y]:
            ws.append([int(y), code, name, cat, bc, bn, rank])
            n += 1
    _finalize(ws, 7, n)

    for y in years:
        sh = wb.create_sheet(y)
        _hdr(sh, ['College Code', 'College Name', 'Seat Category',
                  'Branch Code', 'Branch Name', 'Closing Rank'],
             [12, 52, 12, 12, 52, 13])
        for code, name, cat, bc, bn, rank in data[y]:
            sh.append([code, name, cat, bc, bn, rank])
        _finalize(sh, 6, len(data[y]))

    out = f"COMEDK_Engineering_CutOffs_{years[0]}_{years[-1]}.xlsx"
    wb.save(out)

    print(f'{out}: {n} rows across {len(years)} years')
    for y in years:
        rows = data[y]
        colleges = {r[0] for r in rows}
        branches = {r[3] for r in rows}
        cats = sorted({r[2] for r in rows})
        ranks = [r[5] for r in rows]
        print(f'  {y}: {len(rows):4d} rows  colleges={len(colleges):3d}  '
              f'branches={len(branches):2d}  cats={cats}  rank {min(ranks)}-{max(ranks)}')


if __name__ == '__main__':
    main()
