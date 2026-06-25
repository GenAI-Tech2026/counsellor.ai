#!/usr/bin/env python3
"""
Build the WBJEE 2025 cut-off .xlsx files from the official OR-CR JSON.

Input : source/WBJEE_2025_ORCR_official.json   (see source/fetch_orcr.py)
Output: WBJEE_2025_ENGG_Round1_CutOff.xlsx
        WBJEE_2025_ENGG_Round2_CutOff.xlsx
        WBJEE_2025_ENGG_AllRounds_CutOff.xlsx   (one sheet per round)

Tidy / long layout, mirroring the kcet/ and mhtcet/ datasets: one row per
Institute x Program x Stream x Seat Type x Quota x Category, carrying the
official Opening Rank and Closing Rank for that round.

Re-run:  python3 build_xlsx.py   (from this directory)
"""
import json, os
from openpyxl import Workbook

SRC = "source/WBJEE_2025_ORCR_official.json"

# tidy column order -> source key
COLS = [
    ("Institute", "InstituteName"),
    ("Program", "ProgramName"),
    ("Stream", "StreamName"),
    ("Seat Type", "SeatTypeName"),
    ("Quota", "QuotaName"),
    ("Category", "Category"),
    ("Opening Rank", "OpeningRank"),
    ("Closing Rank", "ClosingRank"),
]
RANK_COLS = {"Opening Rank", "Closing Rank"}


def clean(v):
    return (v or "").strip() if isinstance(v, str) else v


def row_for(rec):
    out = []
    for caption, key in COLS:
        v = clean(rec.get(key))
        if caption in RANK_COLS:
            v = int(v) if str(v).strip().isdigit() else v
        out.append(v)
    return out


def write_sheet(ws, records):
    ws.append([c for c, _ in COLS])
    # stable sort: Institute, Program, Stream, Seat Type, Quota, then Closing Rank
    records = sorted(records, key=lambda r: (
        clean(r.get("InstituteName")) or "",
        clean(r.get("ProgramName")) or "",
        clean(r.get("StreamName")) or "",
        clean(r.get("SeatTypeName")) or "",
        clean(r.get("QuotaName")) or "",
        int(r["ClosingRank"]) if str(r.get("ClosingRank")).strip().isdigit() else 10**9,
    ))
    for rec in records:
        ws.append(row_for(rec))


def main():
    data = json.load(open(SRC))["CounsellingData"]
    rounds = sorted({clean(r["RoundName"]) for r in data})  # Round 1, Round 2

    # per-round workbooks
    for rnd in rounds:
        recs = [r for r in data if clean(r["RoundName"]) == rnd]
        wb = Workbook()
        write_sheet(wb.active, recs)
        wb.active.title = rnd.replace(" ", "")
        fn = f"WBJEE_2025_ENGG_{rnd.replace(' ', '')}_CutOff.xlsx"
        wb.save(fn)
        print(f"{fn:42s} rows={len(recs):5d}  institutes={len({clean(r['InstituteName']) for r in recs})}")

    # combined workbook: one sheet per round
    wb = Workbook()
    wb.remove(wb.active)
    for rnd in rounds:
        recs = [r for r in data if clean(r["RoundName"]) == rnd]
        write_sheet(wb.create_sheet(rnd.replace(" ", "")), recs)
    wb.save("WBJEE_2025_ENGG_AllRounds_CutOff.xlsx")
    print(f"{'WBJEE_2025_ENGG_AllRounds_CutOff.xlsx':42s} rows={len(data):5d}")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
